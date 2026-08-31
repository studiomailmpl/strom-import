"""
Order confirmation parsing — spreadsheets deterministically, PDFs via Claude Vision.

An order confirmation carries the RRP (recommended retail price). Supplier
invoices never do, so this is the pipeline's only source for it.

Dispatch by file type:
  XLSX / CSV  → read the table directly. Deterministic, free, exact.
  PDF         → PyMuPDF for text + page images, then Claude Vision with a
                tool_use schema built for order confirmations.

Google Docs never reach here as such: drive_service.download_file exports them
to PDF, so they arrive as PDF bytes.
"""

import csv
import io
import json
import logging
import re
from dataclasses import dataclass, field, asdict

import anthropic

from app.services.pdf_service import extract_pdf_text, extract_pdf_pages_as_images

logger = logging.getLogger(__name__)

MIME_PDF = "application/pdf"
MIME_XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
MIME_CSV = "text/csv"
GOOGLE_NATIVE_PREFIX = "application/vnd.google-apps."

DEFAULT_MODEL = "claude-sonnet-4-6"

# How far down a sheet to look for the header row before giving up.
HEADER_SCAN_ROWS = 15
# A header row must map at least this many columns to be believed.
MIN_HEADER_MATCHES = 3


@dataclass
class ParsedOrderConfirmationLine:
    style_number: str = ""
    sku: str = ""
    ean: str = ""
    product_name: str = ""
    color_code: str = ""
    color_name: str = ""
    size: str = ""
    quantity: int = 0
    wholesale_price: float | None = None
    rrp: float | None = None

    def as_dict(self) -> dict:
        return asdict(self)


@dataclass
class ParsedOrderConfirmation:
    order_number: str = ""
    vendor: str = ""
    season: str = ""
    currency: str = ""
    lines: list[ParsedOrderConfirmationLine] = field(default_factory=list)
    # Which strategy produced this: "xlsx", "csv" or "ai_vision".
    source: str = ""

    def as_dict(self) -> dict:
        return {
            "order_number": self.order_number,
            "vendor": self.vendor,
            "season": self.season,
            "currency": self.currency,
            "source": self.source,
            "lines": [line.as_dict() for line in self.lines],
        }


class UnsupportedFileType(ValueError):
    """Raised for a file type this parser cannot read."""


# ═══════════════════════════════════════════════
# Column mapping
# ═══════════════════════════════════════════════

# Ordered most specific first — a column is claimed by the first field that
# matches it. "Colour code" must be claimed by color_code before color_name
# sees it, and "Retail price" by rrp before wholesale_price's "price".
COLUMN_SYNONYMS: list[tuple[str, tuple[str, ...]]] = [
    ("ean", ("ean", "barcode", "gtin", "stregkode")),
    ("rrp", ("rrp", "r.r.p", "srp", "msrp", "retail price", "retail",
             "recommended retail", "vejl. udsalg", "vejl udsalg", "udsalgspris",
             "consumer price", "sales price")),
    ("wholesale_price", ("wholesale", "whsl", "wsp", "w/s", "cost price", "cost",
                         "unit price", "net price", "indkobspris", "indkøbspris",
                         "engros", "price")),
    ("color_code", ("colour code", "color code", "col. code", "colorcode",
                    "colourcode", "farvekode", "col code")),
    ("style_number", ("style number", "style no", "style_no", "styleno", "style",
                      "article number", "article no", "article", "artikel",
                      "model", "varenummer")),
    ("sku", ("sku", "item number", "item no", "item", "varenr", "product code")),
    ("size", ("size", "sizes", "str.", "str", "storrelse", "størrelse")),
    ("quantity", ("quantity", "qty", "qnty", "pcs", "units", "antal", "stk")),
    ("color_name", ("colour name", "color name", "colour", "color", "farve")),
    ("product_name", ("product name", "description", "designation", "product",
                      "name", "beskrivelse", "varetekst")),
]

NUMERIC_FIELDS = {"quantity", "wholesale_price", "rrp"}


def _normalise_header(value) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip().lower()


def map_header_row(cells: list) -> dict[int, str]:
    """
    Map column index → field name for a candidate header row.

    Each column is claimed by the first field in COLUMN_SYNONYMS whose synonym
    it contains, and each field is claimed at most once, so a sheet with both
    "Wholesale price" and "Retail price" does not collapse them into one.
    """
    mapping: dict[int, str] = {}
    claimed: set[str] = set()

    for index, cell in enumerate(cells):
        header = _normalise_header(cell)
        if not header:
            continue
        for field_name, synonyms in COLUMN_SYNONYMS:
            if field_name in claimed:
                continue
            if any(synonym in header for synonym in synonyms):
                mapping[index] = field_name
                claimed.add(field_name)
                break

    return mapping


def _find_header(rows: list[list]) -> tuple[int, dict[int, str]]:
    """
    Locate the header row. Confirmations often carry a logo and address block
    above the table, so the header is rarely row 1.

    Returns (row_index, mapping). row_index is -1 when no header was found.
    """
    best_index = -1
    best_mapping: dict[int, str] = {}

    for index, row in enumerate(rows[:HEADER_SCAN_ROWS]):
        mapping = map_header_row(row)
        if len(mapping) > len(best_mapping):
            best_mapping = mapping
            best_index = index

    if len(best_mapping) < MIN_HEADER_MATCHES:
        return -1, {}
    return best_index, best_mapping


def parse_number(value) -> float | None:
    """
    Parse a price cell.

    Handles the European forms confirmations actually use — "1.234,56",
    "48,00", "€ 48.00", "1 234,56" — and returns None when there is no number.
    """
    if value is None:
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)

    text = str(value).strip()
    if not text:
        return None

    # Drop currency symbols, codes and spaces; keep digits, separators, sign.
    text = re.sub(r"[^\d,.\-]", "", text)
    if not text or text in {"-", ".", ","}:
        return None

    if "," in text and "." in text:
        # Whichever separator comes last is the decimal one.
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    elif "," in text:
        # A single comma is a decimal separator unless it groups thousands.
        if re.fullmatch(r"-?\d{1,3}(,\d{3})+", text):
            text = text.replace(",", "")
        else:
            text = text.replace(",", ".")

    try:
        return float(text)
    except ValueError:
        return None


def parse_quantity(value) -> int:
    number = parse_number(value)
    if number is None:
        return 0
    try:
        return int(number)
    except (ValueError, OverflowError):
        return 0


def _cell_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def rows_to_lines(
    rows: list[list], mapping: dict[int, str], start_index: int
) -> list[ParsedOrderConfirmationLine]:
    """Turn data rows below the header into parsed lines."""
    lines: list[ParsedOrderConfirmationLine] = []

    for row in rows[start_index:]:
        if not any(_cell_text(cell) for cell in row):
            continue

        line = ParsedOrderConfirmationLine()
        for index, field_name in mapping.items():
            if index >= len(row):
                continue
            value = row[index]
            if field_name == "quantity":
                line.quantity = parse_quantity(value)
            elif field_name in ("wholesale_price", "rrp"):
                setattr(line, field_name, parse_number(value))
            else:
                setattr(line, field_name, _cell_text(value))

        # A row with no identity is a total, a note or a spacer — not a product.
        if not (line.style_number or line.sku or line.ean or line.product_name):
            continue
        lines.append(line)

    return lines


# ═══════════════════════════════════════════════
# Spreadsheet parsing
# ═══════════════════════════════════════════════

def _header_metadata(rows: list[list], header_index: int) -> dict:
    """
    Scrape order number, season and currency out of the free text above the
    table, where confirmations usually print them.
    """
    text_parts: list[str] = []
    # Everything above the header row. A header on row 0 correctly yields
    # nothing — falling back to a 15-row scan there would read product rows
    # and scrape the order number out of product descriptions.
    for row in rows[:header_index]:
        for cell in row:
            value = _cell_text(cell)
            if value:
                text_parts.append(value)
    blob = " ".join(text_parts)

    meta = {"order_number": "", "season": "", "currency": ""}

    # The captured value must contain a digit, or the heading "ORDER
    # CONFIRMATION" yields order_number="CONFIRMATION" and
    # "Auftragsbestätigung 4711" yields "sbest".
    order_match = re.search(
        r"(?:order|ordre|commande|auftrag)\w*\s*(?:no\.?|nr\.?|number|n[o°])?\s*[:#]?\s*"
        r"(?=[A-Z0-9/-]*\d)([A-Z0-9][A-Z0-9/-]{3,})",
        blob, re.IGNORECASE,
    )
    if order_match:
        meta["order_number"] = order_match.group(1).strip()

    season_match = re.search(
        r"(?:season|saison|collection)\s*[:#]?\s*([A-Za-z][A-Za-z /-]{0,20}?\s?\d{2,4})\b",
        blob, re.IGNORECASE,
    )
    if season_match:
        meta["season"] = season_match.group(1).strip()

    currency_match = re.search(r"\b(EUR|DKK|USD|GBP|SEK|NOK|CHF)\b", blob)
    if currency_match:
        meta["currency"] = currency_match.group(1)
    elif "€" in blob:
        meta["currency"] = "EUR"

    return meta


def _parse_rows(rows: list[list], source: str) -> ParsedOrderConfirmation:
    header_index, mapping = _find_header(rows)
    if header_index < 0:
        logger.warning("No recognisable header row in %s order confirmation", source)
        return ParsedOrderConfirmation(source=source)

    meta = _header_metadata(rows, header_index)
    lines = rows_to_lines(rows, mapping, header_index + 1)

    return ParsedOrderConfirmation(
        order_number=meta["order_number"],
        season=meta["season"],
        currency=meta["currency"],
        lines=lines,
        source=source,
    )


def parse_xlsx(file_bytes: bytes) -> ParsedOrderConfirmation:
    """Parse an .xlsx order confirmation with openpyxl."""
    from openpyxl import load_workbook

    workbook = load_workbook(
        io.BytesIO(file_bytes), read_only=True, data_only=True
    )
    try:
        # Confirmations put the order table on the first sheet; later sheets are
        # size charts and terms.
        sheet = workbook.worksheets[0]
        rows = [list(row) for row in sheet.iter_rows(values_only=True)]
    finally:
        workbook.close()

    return _parse_rows(rows, source="xlsx")


def parse_csv(file_bytes: bytes) -> ParsedOrderConfirmation:
    """Parse a CSV order confirmation, sniffing the delimiter."""
    text = file_bytes.decode("utf-8-sig", errors="replace")

    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        delimiter = dialect.delimiter
    except csv.Error:
        # Sniffing fails on short or irregular files — semicolon is the common
        # default for European exports, but only if it actually appears.
        delimiter = ";" if sample.count(";") > sample.count(",") else ","

    rows = [row for row in csv.reader(io.StringIO(text), delimiter=delimiter)]
    return _parse_rows(rows, source="csv")


# ═══════════════════════════════════════════════
# PDF parsing via Claude Vision
# ═══════════════════════════════════════════════

ORDER_CONFIRMATION_TOOL = {
    "name": "submit_order_confirmation",
    "description": "Submit the parsed contents of a supplier order confirmation.",
    "input_schema": {
        "type": "object",
        "properties": {
            "order_number": {
                "type": "string",
                "description": "Order number from the document header. Empty string if absent.",
            },
            "vendor": {
                "type": "string",
                "description": "Brand/supplier name issuing the confirmation.",
            },
            "season": {
                "type": "string",
                "description": "Season exactly as printed, e.g. 'AW26', 'E26', 'Pre-Spring 2027'. Empty string if absent.",
            },
            "currency": {
                "type": "string",
                "description": "ISO currency code of the prices, e.g. EUR, DKK.",
            },
            "lines": {
                "type": "array",
                "description": "One entry per product line. Split a size grid into one entry per size.",
                "items": {
                    "type": "object",
                    "properties": {
                        "style_number": {"type": "string", "description": "Style/article number as printed"},
                        "sku": {"type": "string", "description": "SKU/item number if shown separately from the style number"},
                        "ean": {"type": "string", "description": "EAN/barcode digits, empty string if absent"},
                        "product_name": {"type": "string", "description": "Product name as printed"},
                        "color_code": {"type": "string", "description": "Colour code, e.g. '3ONXX'"},
                        "color_name": {"type": "string", "description": "Colour name, e.g. 'Dark Navy'"},
                        "size": {"type": "string", "description": "Size label, e.g. 'M', '42'. Use 'One Size' if the product has no sizes."},
                        "quantity": {"type": "integer", "description": "Units ordered for this size"},
                        "wholesale_price": {"type": "number", "description": "Wholesale/cost unit price"},
                        "rrp": {"type": "number", "description": "Recommended retail price per unit"},
                    },
                    "required": ["style_number", "size", "quantity"],
                },
            },
        },
        "required": ["lines"],
    },
}

ORDER_CONFIRMATION_SYSTEM_PROMPT = """Du læser en ORDREBEKRÆFTELSE fra en tøjleverandør og udtrækker den som struktureret data.

En ordrebekræftelse er IKKE en faktura. Den bekræfter hvad butikken har bestilt, og den
indeholder typisk BÅDE en indkøbspris og en vejledende udsalgspris.

RRP ER DET VIGTIGSTE FELT:
"rrp" (recommended retail price) er vejledende udsalgspris — den pris slutkunden skal betale.
Den står ALDRIG på fakturaen, kun her, så den er hele grunden til at dette dokument læses.
Kolonnen hedder typisk: "RRP", "SRP", "MSRP", "Retail", "Retail Price", "Consumer Price",
"Vejl. udsalg", "Udsalgspris", "Sales Price".

FORVEKSL ALDRIG RRP MED INDKØBSPRIS:
"wholesale_price" er hvad butikken betaler. Kolonnen hedder typisk: "Wholesale", "WHSL",
"WSP", "Cost", "Unit Price", "Net", "Indkøbspris", "Engros".
RRP er næsten altid HØJERE end wholesale_price — typisk 2-3 gange. Hvis du kun kan se ÉN
priskolonne, så vurdér ud fra sammenhængen hvilken det er, og lad den anden være tom.
Gæt ALDRIG en RRP ved at gange indkøbsprisen — udelad den hellere.

LINJER:
- Én linje PR. STØRRELSE. Er der et størrelses-grid (S/M/L med antal under hver), så lav
  én linje per størrelse med det antal der står i den kolonne.
- Størrelser med antal 0 eller tom celle skal IKKE med.
- Har produktet ingen størrelser, brug "One Size".
- style_number er varenummeret som det står. Står der både et style-nummer og et separat
  SKU/item-nummer, så udfyld begge; ellers lad "sku" være tom.
- Farvekode og farvenavn er to felter: "3ONXX" er koden, "Dark Navy" er navnet. Står kun
  det ene, udfyld det felt og lad det andet være tomt.
- ean er stregkoden, kun cifre. Tom streng hvis den ikke står.

GENERELT:
- Læs tal PRÆCIST — ciffer for ciffer. Europæiske fakturaer skriver "1.234,56" for
  ét tusind to hundrede og fireogtredive komma seksoghalvtreds.
- Udfyld ALDRIG et felt med et gæt. Tom streng eller udeladt felt er altid bedre end et
  forkert tal.
- Medtag ALLE produktlinjer i dokumentet — også dem der fortsætter på næste side.
- Ignorer totaler, fragt, rabatlinjer og betalingsbetingelser — kun produktlinjer.

Brug submit_order_confirmation-toolen til at returnere resultatet."""


def _build_vision_content(pdf_text: str, pdf_images: list[str]) -> list[dict]:
    parts: list[dict] = []
    for image_b64 in pdf_images:
        parts.append({
            "type": "image",
            "source": {"type": "base64", "media_type": "image/png", "data": image_b64},
        })
    parts.append({
        "type": "text",
        "text": (
            "Udtræk hele denne ordrebekræftelse.\n\n"
            "DOKUMENT-TEKST (supplement til billederne):\n" + pdf_text
        ),
    })
    return parts


def _coerce_ai_lines(raw_lines) -> list[ParsedOrderConfirmationLine]:
    lines: list[ParsedOrderConfirmationLine] = []
    for raw in raw_lines or []:
        if not isinstance(raw, dict):
            continue
        line = ParsedOrderConfirmationLine(
            style_number=str(raw.get("style_number") or "").strip(),
            sku=str(raw.get("sku") or "").strip(),
            ean=str(raw.get("ean") or "").strip(),
            product_name=str(raw.get("product_name") or "").strip(),
            color_code=str(raw.get("color_code") or "").strip(),
            color_name=str(raw.get("color_name") or "").strip(),
            size=str(raw.get("size") or "").strip(),
            quantity=parse_quantity(raw.get("quantity")),
            wholesale_price=parse_number(raw.get("wholesale_price")),
            rrp=parse_number(raw.get("rrp")),
        )
        if not (line.style_number or line.sku or line.product_name):
            continue
        lines.append(line)
    return lines


async def parse_pdf_with_ai(
    file_bytes: bytes,
    *,
    api_key: str,
    model: str = DEFAULT_MODEL,
) -> ParsedOrderConfirmation:
    """Read an order confirmation PDF with Claude Vision."""
    pdf_text = extract_pdf_text(file_bytes)
    pdf_images = extract_pdf_pages_as_images(file_bytes)

    client = anthropic.AsyncAnthropic(api_key=api_key, timeout=120.0)
    message = await client.messages.create(
        model=model,
        max_tokens=16384,
        system=ORDER_CONFIRMATION_SYSTEM_PROMPT,
        messages=[{"role": "user", "content": _build_vision_content(pdf_text, pdf_images)}],
        tools=[ORDER_CONFIRMATION_TOOL],
        tool_choice={"type": "tool", "name": "submit_order_confirmation"},
    )

    payload = None
    for block in message.content:
        if getattr(block, "type", "") == "tool_use" and block.name == "submit_order_confirmation":
            payload = block.input
            break

    if payload is None:
        # tool_choice makes this all but impossible, but a malformed reply must
        # not take the request down.
        logger.warning("No tool_use block in order confirmation response")
        text = "".join(getattr(b, "text", "") for b in message.content)
        match = re.search(r"\{[\s\S]*\}", text)
        if not match:
            return ParsedOrderConfirmation(source="ai_vision")
        try:
            payload = json.loads(match.group(0))
        except json.JSONDecodeError:
            return ParsedOrderConfirmation(source="ai_vision")

    return ParsedOrderConfirmation(
        order_number=str(payload.get("order_number") or "").strip(),
        vendor=str(payload.get("vendor") or "").strip(),
        season=str(payload.get("season") or "").strip(),
        currency=str(payload.get("currency") or "").strip().upper(),
        lines=_coerce_ai_lines(payload.get("lines")),
        source="ai_vision",
    )


# ═══════════════════════════════════════════════
# Entry point
# ═══════════════════════════════════════════════

def _detect_kind(file_name: str, mime_type: str) -> str:
    """Decide how to read a file, trusting the extension over a vague MIME type."""
    name = (file_name or "").lower()
    mime = (mime_type or "").lower()

    if name.endswith(".xlsx") or mime == MIME_XLSX:
        return "xlsx"
    if name.endswith(".csv") or mime == MIME_CSV:
        return "csv"
    if name.endswith(".pdf") or mime == MIME_PDF:
        return "pdf"
    # download_file exports Google Docs to PDF, so the bytes are a PDF even
    # though the file's own type is a Workspace one.
    if mime.startswith(GOOGLE_NATIVE_PREFIX):
        return "pdf"
    return ""


async def parse_order_confirmation(
    file_bytes: bytes,
    file_name: str,
    mime_type: str,
    *,
    api_key: str = "",
    model: str = DEFAULT_MODEL,
) -> ParsedOrderConfirmation:
    """
    Parse an order confirmation into structured data.

    Spreadsheets are read directly — deterministic and free. PDFs go through
    Claude Vision, which needs an api_key.

    Raises UnsupportedFileType for anything else, and ValueError if a PDF is
    given with no API key.
    """
    kind = _detect_kind(file_name, mime_type)

    if kind == "xlsx":
        return parse_xlsx(file_bytes)
    if kind == "csv":
        return parse_csv(file_bytes)
    if kind == "pdf":
        if not api_key:
            raise ValueError("ANTHROPIC_API_KEY is required to parse a PDF order confirmation")
        return await parse_pdf_with_ai(file_bytes, api_key=api_key, model=model)

    raise UnsupportedFileType(
        f"Cannot parse order confirmation {file_name!r} of type {mime_type!r}"
    )
